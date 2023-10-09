import os
import openpyxl


def get_path_to_matrix_file():
    """Get the file path from the user"""
    file_path = input("Path to the file with renaming details: ")
    return file_path


def extract_names_from_file(file_path):
    """Get filenames from Excel and create a dictionary"""
    wb_obj = openpyxl.load_workbook(file_path)

    # load the sheet
    sheet_obj = wb_obj.active
    # get number of rows for interation
    m_row = sheet_obj.max_row
    m_col = sheet_obj.max_column

    source_names = []
    target_names = []
    # loop over the rows and columns and crate two lsits
    for i in range(1, m_row + 1):
        # get cell object by row, column
        for j in range(1, m_col+1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            if cell_obj.column == 1:
                source_names.append(cell_obj.value)
            if cell_obj.column == 2 and cell_obj.value is not None:
                target_names.append(cell_obj.value)
    names_dict = dict(zip(source_names, target_names))
    return names_dict


def rename_files(names_dict):
    """Rename files"""
    target_folder = input("Path to the files you want to rename: ")
    # iterate through all subfolders in the directory
    for root, dirs, files in os.walk(target_folder):
        for file in files:
            # get subfolder name
            subfolder = os.path.join(root.split("\\")[-1])
            # get subfolder/file name
            folder_file_path = os.path.join(subfolder, file)

            # check if subfolder/file in the dictionary keys (source)
            if folder_file_path in names_dict.keys():
                old_name = os.path.join(target_folder, folder_file_path)
                new_name = os.path.join(target_folder, subfolder+"\\"+ names_dict[folder_file_path])
                print(f"Old: {old_name} \n=>New: {new_name}")

                # rename the files
                os.rename(old_name, new_name)


def run():
    """Run program"""
    file_path = get_path_to_matrix_file()
    names_matrix = extract_names_from_file(file_path)
    rename_files(names_matrix)


if __name__ == '__main__':
    run()
